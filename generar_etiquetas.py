# =========================================================
# GENERAR ETIQUETAS DESDE EXCEL
# =========================================================

import os

from openpyxl import load_workbook

from config import EXCEL_PATH, get_output_dir, get_plantillas_map
from etiqueta_renderer import renderizar_etiqueta
from utils import cargar_fuente, limpiar_nombre_archivo


def generar_etiquetas(font_name: str | None = None, progreso_callback=None, excel_path: str | None = None) -> int:

    ruta_excel = excel_path or EXCEL_PATH

    if not os.path.exists(ruta_excel):
        raise FileNotFoundError(f"No se encontró el archivo Excel: {ruta_excel}")

    # =====================================================
    # ABRIR EXCEL
    # =====================================================

    wb = load_workbook(ruta_excel)
    ws = wb.active

    # =====================================================
    # OBTENER HEADERS
    # =====================================================

    headers = [str(cell.value).strip().upper() if cell.value else "" for cell in ws[1]]

    columnas_requeridas = {
        "DESCRIPCION",
        "MARCA",
    }

    faltantes = columnas_requeridas - set(headers)

    if faltantes:
        raise ValueError(
            f"Faltan columnas en el Excel: {', '.join(sorted(faltantes))}"
        )

    # =====================================================
    # ÍNDICES DE COLUMNAS
    # =====================================================

    idx_descripcion = headers.index("DESCRIPCION")
    idx_marca = headers.index("MARCA")

    idx_precio_unidad = (
        headers.index("PRECIO_UNIDAD")
        if "PRECIO_UNIDAD" in headers
        else None
    )

    idx_precio_caja = (
        headers.index("PRECIO_CAJA")
        if "PRECIO_CAJA" in headers
        else None
    )

    # =====================================================
    # CONFIGURACIÓN GENERAL
    # =====================================================

    output_dir = get_output_dir()

    os.makedirs(output_dir, exist_ok=True)

    plantillas = get_plantillas_map()

    font_name = font_name or cargar_fuente()

    total_generadas = 0

    # =====================================================
    # RECORRER FILAS
    # =====================================================

    for fila_numero, fila in enumerate(
        ws.iter_rows(min_row=2, values_only=True),
        start=2,
    ):

        try:

            producto = str(
                fila[idx_descripcion] or ""
            ).upper().strip()

            marca = str(
                fila[idx_marca] or ""
            ).upper().strip()

            precio_unidad = (
                fila[idx_precio_unidad]
                if idx_precio_unidad is not None
                else None
            )

            precio_caja = (
                fila[idx_precio_caja]
                if idx_precio_caja is not None
                else None
            )

        except Exception as exc:

            mensaje = f"[ERROR] Fila {fila_numero}: error leyendo datos: {exc}"

            print(mensaje)

            if progreso_callback:
                progreso_callback(mensaje)

            continue

        # =================================================
        # VALIDAR MARCA
        # =================================================

        if marca not in plantillas:

            mensaje = f"[ERROR] Fila {fila_numero}: plantilla/marca no encontrada: {marca}. Debe coincidir con el nombre del archivo de la plantilla sin extensión."

            print(mensaje)

            if progreso_callback:
                progreso_callback(mensaje)

            continue

        # =================================================
        # GENERAR IMAGEN
        # =================================================

        try:

            imagen = renderizar_etiqueta(
                producto=producto,
                marca=marca,
                precio_unidad=precio_unidad,
                precio_caja=precio_caja,
                font_name=font_name,
            )

        except Exception as exc:

            mensaje = f"[ERROR] Fila {fila_numero}: {exc}"

            print(mensaje)

            if progreso_callback:
                progreso_callback(mensaje)

            continue

        # =================================================
        # GUARDAR PNG
        # =================================================

        nombre_archivo = limpiar_nombre_archivo(producto)

        output_path = os.path.join(
            output_dir,
            f"{nombre_archivo}.png"
        )

        imagen.save(output_path)

        total_generadas += 1

        mensaje = f"[OK] Generado: {nombre_archivo}.png"

        print(mensaje)

        if progreso_callback:
            progreso_callback(mensaje)

    # =====================================================
    # FIN PROCESO
    # =====================================================

    mensaje = f"Proceso terminado. Etiquetas generadas: {total_generadas}"

    print(mensaje)

    if progreso_callback:
        progreso_callback(mensaje)

    return total_generadas